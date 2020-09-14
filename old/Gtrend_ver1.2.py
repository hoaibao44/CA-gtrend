import json
import sys
import time
from pprint import pprint
from datetime import date, datetime, timedelta
from pytrends.request import TrendReq
from time import sleep


import pandas as pd
import requests
from requests.adapters import HTTPAdapter
from requests.packages.urllib3.util.retry import Retry
import openpyxl
from openpyxl import Workbook, load_workbook

from __future__ import print_function
import pickle
import os.path
import requests
from googleapiclient.discovery import build
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request

TEST = '159362404'
CW_ID = '173996740'
cl_name = 'FOD'

def sendMessage(room_id, text_message_as_string):
    APIKEY = '66d7468c0232a7a8d0c24d8ef8c2a71c'
    ENDPOINT = 'https://api.chatwork.com/v2'
    room_id = room_id
    bodytext = text_message_as_string
    post_message_url = '{}/rooms/{}/messages'.format(ENDPOINT, room_id)
         
    headers = { 'X-ChatWorkToken': APIKEY }
    #print(str(bodytext))
    params = { 'body': str(bodytext) }
         
    resp = requests.post(post_message_url,
                         headers=headers,
                         params=params)
    print(resp.encoding)
    #print(resp.url)
    return pprint(resp.content)

def sendFile(room_id,file_name,file_path):
    APIKEY = '66d7468c0232a7a8d0c24d8ef8c2a71c'
    ENDPOINT = 'https://api.chatwork.com/v2'
    room_id = room_id
    file_path = file_path
    file_name = file_name
    post_message_url = '{}/rooms/{}/files'.format(ENDPOINT, room_id)
         
    headers = { 'X-ChatWorkToken': APIKEY }

    files  = {'file': (file_name,open(file_path, 'rb'),'application/vnd.ms-excel')}
         
    resp = requests.post(post_message_url,headers=headers,files=files)
    #print(resp.url)
    return pprint(resp.content)




#----MAIN PROCESS----

min_value = 5000
process_start = datetime.now()

#read KW_list from file
file_path = os.getcwd()+'\\対象KW.xlsx'
wb = load_workbook(file_path)
ws = wb.active
KW_list = []
for row in ws:
	for cell in row:
		if cell.value != '対象KW一覧':
			KW_list.append(cell.value)
print('Start processing for: ' + str(len(KW_list)) + ' KWs')


CW_ALL_msg =''
#pytrends = TrendReq(hl='ja-JP', tz=360,proxies=['https://10.141.64.176:3128','https://172.19.119.54:3128'], retries=3)
pytrends = TrendReq(hl='ja-JP', tz=360, retries=3)


#output file
now = datetime.now().strftime("%Y/%m/%d %H:%M:%S")
now_time = datetime.now().strftime("%H:%M:%S")
cl_title = ['実行時点','KW','クエリ','状況']
rows_data =[]
rows_data.append(cl_title)


for inKW in KW_list:
    try:
      print('●[' +str(KW_list.index(inKW)+1) + ' / ' +str(len(KW_list))+ '] '+ inKW + ' is processing!')
      sleep(10)
      #rising_df = get_query(inKW)
      kw_list = [inKW]
      pytrends.build_payload(kw_list, timeframe='now 1-H', geo='JP')
      df = pytrends.related_queries()
      rising_df = df[inKW]['rising']
      #print(rising_df)
      if rising_df is not None:
        i=0
        CW_msg_query = ''
        CW_msg=''
        while True: 
          if rising_df.loc[i]['value'] >= min_value and rising_df.loc[i]['value'] < 5000:
            #print(rising_df.loc[i]['query'])
            CW_msg_query =  CW_msg_query +'└\t'+ rising_df.loc[i]['query'] +'（'+ str(rising_df.loc[i]['value']) +'%増加）\n'
            rows_data.append([now,inKW,rising_df.loc[i]['query'],'（'+ str(rising_df.loc[i]['value']) +'%増加）'])
          elif rising_df.loc[i]['value'] >= 5000:
            CW_msg_query =  CW_msg_query +'└\t'+ rising_df.loc[i]['query'] +'（急激増加）\n'
            rows_data.append([now,inKW,rising_df.loc[i]['query'],'（急激増加）'])  
          i=i+1
          if(i == rising_df['value'].count()):  
                break
              
        if CW_msg_query !='':
           CW_msg = '▼Rising query for KW:  ' + inKW + ' \n' +   CW_msg_query
           print(CW_msg)
           CW_ALL_msg = CW_ALL_msg + CW_msg +'\n'
                   
    except KeyError:
      print(inKW + ' has some kind of errors!')

sendMessage(CW_ID,"[To:3656340][To:1376387][info][title]Google Trend's Rising Query Info: "+str(now) +"[/title]"+ CW_ALL_msg +'[/info]')
print(CW_ALL_msg)

#write file:
file_path = os.getcwd()+'\\' +cl_name +'_'+ (date.today() - timedelta(days=1)).strftime('%Y%m') + '月_rising_query_info.xlsx'

#check if file has info
if os.path.isfile(file_path):
    has_info = True
    print('already has file')
else:
    has_info = False
    print('first time')

if has_info == False: #first run time
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title =  (date.today() - timedelta(days=1)).strftime('%Y%m') + '月'
    for row in rows_data:
         ws.append(row)
    wb.save(file_path)
else: #not the first run time
    wb = load_workbook(file_path)
    ws = wb.active
    max_row = ws.max_row
    #print(max_row)
    for row in rows_data:
        if row[0] != '実行時点':   #ignore first row             
            cl = 1
            for cell in row:
                ws.cell(max_row+1,cl,cell)
                cl = cl+ 1
            max_row = max_row +1
        
    wb.save(file_path)

#if run at 10AM then sent yesterday info to CW:
print(now_time)
if '09:59:00' < now_time <'10:45:00':
    send_file_CW = True
else:
    send_file_CW = False

print('send_file_CW = ' + str(send_file_CW))

if send_file_CW == True:
    sendMessage(CW_ID,'[To:3656340][To:1376387][info]'+str((date.today() - timedelta(days=1)).strftime('%Y/%m/%d'))+" 's Monthly Query Info[/info]")
    sendFile(CW_ID,cl_name +'_'+(date.today() - timedelta(days=1)).strftime('%Y%m') + '月_rising_query_info.xlsx',file_path)

process_duration = (datetime.now() - process_start).seconds

print('※DONE IN※', process_duration, ' seconds' )
    
      
